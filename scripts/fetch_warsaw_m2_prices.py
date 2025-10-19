#!/usr/bin/env python3
"""
Warsaw Real Estate Price Data Downloader and Processor

This script fetches quarterly housing price data from the National Bank of Poland (NBP),
extracts Warsaw m2 prices, interpolates to monthly granularity, and converts prices to
gold equivalents using NBP gold price data.

NBP Data Source: https://static.nbp.pl/dane/rynek-nieruchomosci/ceny_mieszkan.xlsx
- Quarterly data available from Q3 2006 onwards
- Contains average prices per m2 for major Polish cities including Warsaw

Usage:
    python fetch_warsaw_m2_prices.py [--output ../data/warsaw-m2-prices-monthly.json] [--verbose]
"""

import json
import requests
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
import argparse
import sys
from io import BytesIO

# Try to import openpyxl for Excel parsing
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class WarsawM2PriceFetcher:
    """Fetches and processes Warsaw real estate price data from NBP."""
    
    NBP_EXCEL_URL = "https://static.nbp.pl/dane/rynek-nieruchomosci/ceny_mieszkan.xlsx"
    GOLD_PRICES_FILE = Path(__file__).parent.parent / 'data' / 'nbp-gold-prices-monthly.json'
    
    def __init__(self, verbose: bool = False):
        self.verbose = verbose
        self.session = requests.Session()
        self.session.headers.update({'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})
    
    def log(self, message: str):
        """Print message if verbose mode is enabled."""
        if self.verbose:
            print(f"[INFO] {message}")
    
    def fetch_excel_data(self) -> bytes:
        """
        Fetch the NBP Excel file with housing prices.
        
        Returns:
            Raw bytes of the Excel file
        """
        self.log(f"Downloading NBP housing prices data from {self.NBP_EXCEL_URL}")
        
        try:
            response = self.session.get(self.NBP_EXCEL_URL, timeout=30)
            response.raise_for_status()
            self.log(f"  → Downloaded {len(response.content)} bytes")
            return response.content
        except requests.exceptions.RequestException as e:
            print(f"[ERROR] Failed to download NBP data: {e}", file=sys.stderr)
            raise
    
    def extract_warsaw_quarterly_prices(self, excel_data: bytes) -> List[Dict]:
        """
        Extract Warsaw m2 prices from the NBP Excel file.
        
        Args:
            excel_data: Raw bytes of the Excel file
            
        Returns:
            List of dicts with 'year', 'quarter', and 'priceM2' keys
        """
        if not OPENPYXL_AVAILABLE:
            print("[ERROR] openpyxl not installed. Install it with: pip install openpyxl", file=sys.stderr)
            raise ImportError("openpyxl is required to parse Excel files")
        
        self.log("Parsing Excel file...")
        
        workbook = openpyxl.load_workbook(BytesIO(excel_data))
        self.log(f"  → Available sheets: {workbook.sheetnames}")
        
        # Try to find the main data sheet (usually the first one)
        sheet = workbook.active
        self.log(f"  → Using sheet: {sheet.title}")
        
        warsaw_prices = []
        
        # Find header row and Warsaw column
        warsaw_col = None
        header_row = None
        
        # Scan for headers
        for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
            if any(cell and 'warsza' in str(cell).lower() for cell in row):
                header_row = row_idx
                warsaw_col = next(
                    (idx + 1 for idx, cell in enumerate(row) if cell and 'warsza' in str(cell).lower()),
                    None
                )
                self.log(f"  → Found Warsaw column at row {header_row}, column {warsaw_col}")
                break
        
        if not warsaw_col:
            print("[ERROR] Could not find Warsaw column in Excel file", file=sys.stderr)
            raise ValueError("Warsaw column not found in Excel data")
        
        # Extract data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            # Expect: Period (e.g., "Q1 2023"), Warsaw price, ...
            period_cell = row[0]
            price_cell = row[warsaw_col - 1]  # Adjust for 0-based indexing
            
            if not period_cell or not price_cell:
                continue
            
            # Parse period string (e.g., "Q1 2023" or "I kw. 2023" or "1 kw. 2023")
            period_str = str(period_cell).strip()
            
            # Try to extract year and quarter
            year, quarter = self._parse_period(period_str)
            
            if year is None or quarter is None:
                self.log(f"  ⚠ Skipping unrecognized period: {period_str}")
                continue
            
            try:
                price = float(price_cell)
                warsaw_prices.append({
                    'year': year,
                    'quarter': quarter,
                    'priceM2': price
                })
            except (ValueError, TypeError):
                self.log(f"  ⚠ Skipping invalid price for {period_str}: {price_cell}")
                continue
        
        self.log(f"  → Extracted {len(warsaw_prices)} quarterly data points for Warsaw")
        return warsaw_prices
    
    def _parse_period(self, period_str: str) -> Tuple[int, int]:
        """
        Parse period string to extract year and quarter.
        
        Supports formats:
        - "I 2006", "II 2006", "III 2006", "IV 2006" (Roman numeral + space + year)
        - "Q1 2023"
        - "1 kw. 2023"
        - "I kw. 2023"
        - "1 kw 2023"
        
        Returns:
            Tuple of (year, quarter) or (None, None) if parsing failed
        """
        import re
        
        period_lower = period_str.lower()
        
        # Roman numerals to numbers
        roman_to_int = {'i': 1, 'ii': 2, 'iii': 3, 'iv': 4}
        
        year = None
        quarter = None
        
        # Try to extract year (4-digit year)
        year_match = re.search(r'\b(20\d{2})\b', period_str)
        if year_match:
            year = int(year_match.group(1))
        
        # Try Roman numeral format: "I 2006", "II 2006", "III 2006", "IV 2006"
        if quarter is None:
            roman_match = re.search(r'\b([ivIV]+)\s+(?:20\d{2})', period_str)
            if roman_match:
                q_str = roman_match.group(1).lower()
                if q_str in roman_to_int:
                    quarter = roman_to_int[q_str]
        
        # Try Q format: "Q1", "Q2", etc.
        if quarter is None and 'q' in period_lower:
            q_match = re.search(r'q([1-4])', period_lower)
            if q_match:
                quarter = int(q_match.group(1))
        
        # Polish format with "kw": "1 kw.", "I kw.", etc.
        if quarter is None and 'kw' in period_lower:
            # Look for number before "kw"
            kw_match = re.search(r'([1-4ivVI]+)\s*kw', period_lower)
            if kw_match:
                q_str = kw_match.group(1).lower()
                if q_str in roman_to_int:
                    quarter = roman_to_int[q_str]
                else:
                    try:
                        quarter = int(q_str)
                    except ValueError:
                        pass
        
        return year, quarter
    
    def interpolate_quarterly_to_monthly(self, quarterly_prices: List[Dict]) -> List[Dict]:
        """
        Interpolate quarterly prices to monthly granularity using linear interpolation.
        
        For each quarter, assigns the same price to all 3 months of that quarter,
        then interpolates between quarter boundaries to smooth transitions.
        
        Args:
            quarterly_prices: List of dicts with 'year', 'quarter', 'priceM2' keys
            
        Returns:
            List of dicts with 'year', 'month', and 'priceM2' keys
        """
        self.log("Interpolating quarterly data to monthly...")
        
        if not quarterly_prices:
            return []
        
        # Sort by year and quarter
        sorted_prices = sorted(quarterly_prices, key=lambda x: (x['year'], x['quarter']))
        
        monthly_prices = {}
        
        # First pass: assign quarter prices to all months in that quarter
        for entry in sorted_prices:
            year = entry['year']
            quarter = entry['quarter']
            price = entry['priceM2']
            
            # Calculate starting month (0-based quarter index to months)
            start_month = (quarter - 1) * 3 + 1
            
            for month_offset in range(3):
                month = start_month + month_offset
                if month <= 12:
                    key = (year, month)
                    monthly_prices[key] = price
        
        # Second pass: interpolate between available quarters for smoother transitions
        for i, entry in enumerate(sorted_prices[:-1]):
            year1 = entry['year']
            quarter1 = entry['quarter']
            price1 = entry['priceM2']
            
            next_entry = sorted_prices[i + 1]
            year2 = next_entry['year']
            quarter2 = next_entry['quarter']
            price2 = next_entry['priceM2']
            
            # Convert to month indices for calculation
            month1 = (year1 - 2006) * 12 + (quarter1 - 1) * 3 + 2  # Mid-quarter month
            month2 = (year2 - 2006) * 12 + (quarter2 - 1) * 3 + 2
            
            # Interpolate months between quarters
            if month2 > month1 + 3:
                num_steps = month2 - month1
                for step in range(1, num_steps):
                    interpolated_price = price1 + (price2 - price1) * (step / num_steps)
                    
                    # Calculate the year and month
                    total_months = month1 + step
                    calc_year = 2006 + (total_months - 1) // 12
                    calc_month = ((total_months - 1) % 12) + 1
                    
                    # Only update if within expected range
                    if calc_month >= 1 and calc_month <= 12:
                        key = (calc_year, calc_month)
                        if key not in monthly_prices:
                            monthly_prices[key] = round(interpolated_price, 2)
        
        # Convert to list format sorted by date
        result = []
        for (year, month), price in sorted(monthly_prices.items()):
            result.append({
                'year': year,
                'month': month,
                'priceM2_pln': round(price, 2)
            })
        
        self.log(f"  → Generated {len(result)} monthly data points")
        return result
    
    def load_gold_prices(self) -> Dict[Tuple[int, int], float]:
        """
        Load monthly gold prices from JSON file.
        
        Returns:
            Dict mapping (year, month) to gold price in PLN per gram
        """
        self.log(f"Loading gold prices from {self.GOLD_PRICES_FILE}")
        
        try:
            with open(self.GOLD_PRICES_FILE, 'r', encoding='utf-8') as f:
                gold_data = json.load(f)
            
            gold_prices = {}
            for entry in gold_data:
                key = (entry['year'], entry['month'])
                gold_prices[key] = entry['price']
            
            self.log(f"  → Loaded {len(gold_prices)} monthly gold prices")
            return gold_prices
        except FileNotFoundError:
            print(f"[ERROR] Gold prices file not found: {self.GOLD_PRICES_FILE}", file=sys.stderr)
            raise
        except (json.JSONDecodeError, KeyError) as e:
            print(f"[ERROR] Invalid gold prices file format: {e}", file=sys.stderr)
            raise
    
    def convert_to_gold_equivalent(self, monthly_prices: List[Dict], gold_prices: Dict) -> List[Dict]:
        """
        Convert PLN prices to gold equivalent (grams of gold).
        
        Formula: priceM2_gold = priceM2_pln / gold_price_per_gram
        
        Args:
            monthly_prices: List of dicts with 'year', 'month', 'priceM2_pln' keys
            gold_prices: Dict mapping (year, month) to gold price in PLN per gram
            
        Returns:
            List of dicts with added 'priceM2_gold' key (grams of 1g gold equivalent)
        """
        self.log("Converting prices from PLN to gold equivalent...")
        
        result = []
        missing_months = set()
        
        for entry in monthly_prices:
            year = entry['year']
            month = entry['month']
            price_pln = entry['priceM2_pln']
            
            key = (year, month)
            
            if key in gold_prices:
                gold_price = gold_prices[key]
                price_gold = price_pln / gold_price
                result.append({
                    'year': year,
                    'month': month,
                    'priceM2_pln': price_pln,
                    'priceM2_gold': round(price_gold, 2)
                })
            else:
                missing_months.add(key)
                result.append({
                    'year': year,
                    'month': month,
                    'priceM2_pln': price_pln,
                    'priceM2_gold': None  # Unable to convert
                })
        
        if missing_months:
            self.log(f"  ⚠ Warning: {len(missing_months)} months missing gold price data")
            for year, month in sorted(missing_months)[:5]:  # Show first 5
                self.log(f"     {year}-{month:02d}")
            if len(missing_months) > 5:
                self.log(f"     ... and {len(missing_months) - 5} more")
        
        self.log(f"  → Converted {len(result)} entries to gold equivalent")
        return result
    
    def save_json(self, data: List[Dict], filepath: Path, pretty: bool = True):
        """
        Save data to JSON file.
        
        Args:
            data: Data to save
            filepath: Output file path
            pretty: Whether to pretty-print JSON
        """
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2 if pretty else None, ensure_ascii=False)
        
        self.log(f"Saved {len(data)} entries to {filepath}")


def main():
    parser = argparse.ArgumentParser(
        description='Fetch Warsaw m2 prices from NBP, interpolate to monthly, and convert to gold equivalent.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Fetch and process data (default)
  python fetch_warsaw_m2_prices.py

  # Custom output file
  python fetch_warsaw_m2_prices.py --output custom_output.json

  # Verbose output
  python fetch_warsaw_m2_prices.py -v
        """
    )
    
    parser.add_argument(
        '--output',
        type=Path,
        default=Path(__file__).parent.parent / 'data' / 'warsaw-m2-prices-monthly.json',
        help='Output JSON file path (default: ../data/warsaw-m2-prices-monthly.json)'
    )
    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Enable verbose output'
    )
    
    args = parser.parse_args()
    
    print(f"🏠 Warsaw M2 Price Processor")
    print(f"{'=' * 60}")
    
    fetcher = WarsawM2PriceFetcher(verbose=args.verbose)
    
    try:
        # Step 1: Fetch Excel data from NBP
        print(f"\n📥 Fetching Data:")
        print(f"{'=' * 60}")
        excel_data = fetcher.fetch_excel_data()
        
        # Step 2: Extract Warsaw quarterly prices
        print(f"\n📊 Extracting Warsaw Data:")
        print(f"{'=' * 60}")
        quarterly_prices = fetcher.extract_warsaw_quarterly_prices(excel_data)
        
        if not quarterly_prices:
            print("[ERROR] No quarterly data extracted for Warsaw", file=sys.stderr)
            return 1
        
        print(f"✅ Extracted {len(quarterly_prices)} quarterly data points")
        print(f"   Date range: Q{quarterly_prices[0]['quarter']} {quarterly_prices[0]['year']} to "
              f"Q{quarterly_prices[-1]['quarter']} {quarterly_prices[-1]['year']}")
        print(f"   Price range: {min(p['priceM2'] for p in quarterly_prices):.2f} - "
              f"{max(p['priceM2'] for p in quarterly_prices):.2f} PLN/m²")
        
        # Step 3: Interpolate to monthly
        print(f"\n📈 Processing Data:")
        print(f"{'=' * 60}")
        monthly_prices = fetcher.interpolate_quarterly_to_monthly(quarterly_prices)
        
        if not monthly_prices:
            print("[ERROR] Failed to interpolate data to monthly", file=sys.stderr)
            return 1
        
        print(f"✅ Interpolated to {len(monthly_prices)} monthly data points")
        
        # Step 4: Load gold prices and convert
        print(f"\n💰 Converting to Gold Equivalent:")
        print(f"{'=' * 60}")
        gold_prices = fetcher.load_gold_prices()
        
        monthly_prices_with_gold = fetcher.convert_to_gold_equivalent(monthly_prices, gold_prices)
        
        # Filter out entries without gold conversion (shouldn't happen with full data)
        complete_entries = [e for e in monthly_prices_with_gold if e['priceM2_gold'] is not None]
        print(f"✅ Converted {len(complete_entries)} entries to gold equivalent")
        print(f"   Price range: {min(e['priceM2_gold'] for e in complete_entries):.2f} - "
              f"{max(e['priceM2_gold'] for e in complete_entries):.2f} g/m²")
        
        # Step 5: Save results
        print(f"\n💾 Saving Results:")
        print(f"{'=' * 60}")
        fetcher.save_json(monthly_prices_with_gold, args.output)
        
        print(f"\n📁 Output: {args.output.resolve()}")
        print(f"✨ Done! Generated monthly data with PLN and gold equivalent prices.")
        
        return 0
        
    except Exception as e:
        print(f"\n[ERROR] {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
